import requests
import jsonpath
import datetime
import openpyxl
import os.path
import logging

def shopGrade(session,shopId,acctId,token, cookie): # 7日数据，确定美团外卖店铺等级
    
    url = "https://waimaieapp.meituan.com/bizdata/overviewV4/trend"        
    beginDate = (datetime.datetime.now() + datetime.timedelta(days=-7)).strftime("%Y%m%d")        
    endDate = (datetime.datetime.now() + datetime.timedelta(days=-1)).strftime("%Y%m%d")        
    params = {                
            "analysisType": 21,     "indicatorType": 0,     "durationType": 2,        "businessCircleType": 1, # 1 stand for top , 0 for average           
            "beginDate": beginDate,     "endDate": endDate,      "token":token,  "wmPoiId":shopId,   "acctId": acctId,   "appType": 3          
            }
    headers = {                
            "Accept":"application/json, text/plain, */*",            
            "Accept-Encoding":"gzip, deflate, br",            
            "Accept-Language":"zh-CN,zh;q=0.8",            
            "Connection":"keep-alive",            
            "Cookie":cookie + str(shopId),            
            "Host":"waimaieapp.meituan.com",            
            "Referer":"https://waimaieapp.meituan.com/igate/bizdata/business",            
            "User-Agent":"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4103.116 Safari/537.36"
        }
    try:
        requests.packages.urllib3.disable_warnings()
        res = session.get(url,headers=headers,params=params,verify=False)            
        data = res.json()          
        base_value = round(sum(jsonpath.jsonpath(data,"$..data.trendChart.base")[0])) # 本店周实收       
        top_value = round(sum(jsonpath.jsonpath(data,"$..data.trendChart.businessCircle")[0])) # top 10 数据
        
        params["businessCircleType"]=0    
        requests.packages.urllib3.disable_warnings()
        res2 = session.get(url,headers=headers,params=params,verify=False)   
        data2 = res2.json() 
        avr_value = round(sum(jsonpath.jsonpath(data2,"$..data.trendChart.businessCircle")[0]))  # 商圈均值
            
        if top_value == 0:                   
            return [endDate,"-",0,0]                
        elif base_value >= top_value: 
            flag = "A"          
        elif base_value <= avr_value:            
            flag = "C"
        elif abs(top_value - base_value) >= abs(base_value-avr_value):            
            flag = "BC"           
        else:                   
             flag = "BA" 


        url3 = "https://waimaieapp.meituan.com/bizdata/overviewV4/history"       
        params3 = { "durationType":2,  "terminalType":2,  "beginDate":beginDate,     "endDate":endDate,     "token":token,     "wmPoiId":shopId,   "acctId":acctId, "appType":3 }              
        requests.packages.urllib3.disable_warnings()              
        res3 = session.get(url3,headers=headers,params=params3,verify=False)
        data3 = res3.json()  
        delta = jsonpath.jsonpath(data3,"$..data.business.settleAmountDelta")[0] # 趋势    
        settle = jsonpath.jsonpath(data3,"$..data.business.settleAmount")[0]

        if settle == delta: return [endDate,flag,base_value,top_value] # 上一周未营业或两周期等值，则flag无附加                
        if delta>0 and delta/(settle-delta)>0.05:             
            flag +=  "+"                  
        elif delta<0 and abs(delta)/(settle-delta)>0.05: # 上下波动在5%以内，都是不变     
            flag +=  "-" 
        return [endDate, base_value, top_value, flag]    
    except:   
            return False


def getScore(session, shopId, token, acctId, bsid, cookie):

    url = "https://waimaieapp.meituan.com/gw/customer/comment/scores?"    
    
    headers = {           
        "Accept": "application/json, text/plain, */*",        
        "Accept-Encoding": "gzip, deflate, br",        
        "Accept-Language":"zh-CN,zh;q=0.9",        
        "Connection":"keep-alive",    
        "Cookie":cookie + str(shopId),        
        "Host": "waimaieapp.meituan.com",        
        "Origin": "https://waimaieapp.meituan.com",        
        "Referer": "https://waimaieapp.meituan.com",        
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4324.182 Safari/537.36"  
        }

    payload = {"ignoreSetRouterProxy":"true",  "token":token,   "acctId":acctId,    "wmPoiId":shopId,  "bsid":bsid,  "appType":3}
    
    requests.packages.urllib3.disable_warnings()   
    res = session.post(url,headers=headers,data=payload,verify=False)    
    data = res.json()
    score = jsonpath.jsonpath(data,"$..data.poiScore")

    if score == False:    
        return 0 
    else:
        return score[0]



def dayinfo(session,shopId,acctId,token, cookie):

    url = "https://waimaieapp.meituan.com/bizdata/overviewV4/history"
    yesterday =  (datetime.date.today() + datetime.timedelta(days=-1)).strftime("%Y%m%d")   
    params = {"beginDate":yesterday,   "endDate":yesterday,   "durationType":1,    "terminalType":2,    "token":token,  "wmPoiId":shopId,   "acctId":acctId,   "appType": 3}
    
    headers = {
        "Accept":"application/json, text/plain, */*",
        "Accept-Encoding":"gzip, deflate, br",
        "Accept-Language":"zh-CN,zh;q=0.8",
        "Connection":"keep-alive",
        "Cookie":cookie + str(shopId),
        "Host":"waimaieapp.meituan.com",
        "Referer":"https://waimaieapp.meituan.com/igate/bizdata/business",
        "User-Agent":"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4103.116 Safari/537.36"
    }
    requests.packages.urllib3.disable_warnings()
    res = session.get(url, headers = headers, params = params, verify = False)
    
    try:    
            data = res.json()     
            settle = round(jsonpath.jsonpath(data,"$..data.business.settleAmount")[0]) # 实收
            orderCnt = jsonpath.jsonpath(data,"$..data.business.orderCnt")[0] # 订单量     
            avrPrice = round(jsonpath.jsonpath(data,"$..data.business.avgPrice")[0])# 单均
            expoNum = jsonpath.jsonpath(data,"$..data.flow.exposureNum")[0] # 曝光人数               
            visitRate = round(jsonpath.jsonpath(data,"$..data.flow.visitRate")[0],4) # 进店率            
            orderRate = round(jsonpath.jsonpath(data,"$..data.flow.orderRate")[0],4) # 下单率
            return [settle, avrPrice, orderCnt, expoNum, visitRate, orderRate]   
    except:  
         return False
    

    

def cost_effect(session,shopId,acctId,token, cookie): # 活动占比、佣金占比、到手率
    
    url = "https://waimaieapp.meituan.com/bizdata/businessStatisticsV4/incomeDetail"
    yesterday = (datetime.date.today() + datetime.timedelta(days=-1)).strftime("%Y%m%d")
    params = {"beginDate":yesterday,"endDate":yesterday,"durationType":1,"token":token,"wmPoiId":shopId,"acctId":acctId,"appType":3 }
    
    headers = {       
        "Accept":"application/json, text/plain, */*",        
        "Accept-Encoding":"gzip, deflate, br",        
        "Accept-Language":"zh-CN,zh;q=0.8",        
        "Connection":"keep-alive",        
        "Cookie":cookie + str(shopId),       
        "Host":"waimaieapp.meituan.com",        
        "Referer":"https://waimaieapp.meituan.com/igate/bizdata/business",      
        "User-Agent":"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36"    
    }
    requests.packages.urllib3.disable_warnings()
    res = requests.get(url=url,headers=headers,params=params,verify=False)
    
    try:
        data = res.json()
        turnover = jsonpath.jsonpath(data,"$..foodPrice")[0]+jsonpath.jsonpath(data,"$..packFee")[0]
        if turnover == 0: # 昨日未营业                
           return [0, 0, 0]      
        promRate = round(jsonpath.jsonpath(data,"$..actExp")[0]/turnover, 4) 
        platRate = round(jsonpath.jsonpath(data,"$..platFee")[0]/(turnover - jsonpath.jsonpath(data,"$..actExp")[0]), 4)  
        settle = turnover - jsonpath.jsonpath(data,"$..actExp")[0] - jsonpath.jsonpath(data,"$..platFee")[0] - jsonpath.jsonpath(data,"$..otherFlow")[0]   
        settleRate = round(settle/turnover,4)   
        return [promRate, settleRate, platRate]
    except:    
          return False


def retry(func,*args):
    result = func(*args)
    count = 0  
    while result == False and count < 5:
        result = retry(func,*args)
        count += 1 
    return result

    
def main(token_dic):

    cookie = token_dic["cookie"]
    shop_dic = token_dic["mt_shop_dic"]
    acctId = token_dic["acctId"]
    token = token_dic["token"]
    bsid = token_dic["bsid"]
    data_dic = {}
    session = requests.Session()
  
    for shopId in shop_dic:
        a = retry(shopGrade, session, shopId, acctId, token, cookie)
        b = retry(dayinfo, session, shopId, acctId, token, cookie)  
        c = retry(cost_effect, session, shopId, acctId, token, cookie)  
        if a == False or b == False or c == False :
            continue
        score = getScore(session, shopId, token, acctId, bsid, cookie)
        a.append(score)
        data_dic[shopId] = a + b + c


    path = "美团日记录.xlsx"
    if os.path.exists(path) == False:
        wb = openpyxl.load_workbook(r".\templete\日报模板.xlsx")
        ws = wb["美团"]
        for shopName in shop_dic.values():
            _ = wb.copy_worksheet(ws)
            _.title = shopName

        wb.remove(ws)           
        wb.remove(wb["饿了么"])
        wb.save(path)
    
    wb = openpyxl.load_workbook(path)     
    up_font = openpyxl.styles.Font(color="FF0000")   # 格式美化
    down_font = openpyxl.styles.Font(color="00FFFF")

    for shopId in shop_dic:
        shop = shop_dic[shopId]
        wb[shop].append(data_dic[shopId])        
        max_row = wb[shop].max_row
        while wb[shop].cell(max_row,1).value is None:
            max_row -= 1

        for col in range(1,15):
            wb[shop].cell(max_row,col).alignment = openpyxl.styles.Alignment("center")
        
        for col in range(10,15): # 双转、自营销、到手率显示为百分比
            wb[shop].cell(max_row,col).number_format = "0.0%"

        if max_row >= 9 :
            for col in range(6,15):
                if wb[shop].cell(max_row,col).value >= wb[shop].cell(max_row-7,col).value * 1.05 :
                    wb[shop].cell(max_row,col).font = up_font
                    continue
                if wb[shop].cell(max_row,col).value <= wb[shop].cell(max_row-7,col).value * 0.95 :
                    wb[shop].cell(max_row,col).font = down_font 
                                        
    wb.save(path)
