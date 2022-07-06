import requests
import datetime
import jsonpath
import os
import openpyxl

def fetch(shopId, token, acctId, cookie):

    data_date = (datetime.date.today() + datetime.timedelta(days=-1)).strftime("%Y%m%d")    
    params = {
        "tabType":1,      "durationType":1,     "beginDate":data_date,      "endDate":data_date,        
        "token":token,    "wmPoiId":shopId,      "acctId":acctId,        "appType":3
    }
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Cookie": cookie + str(shopId),
        "Referer": "https://waimaieapp.meituan.com/igate/bizdata/flow",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36"
    }
    res = requests.get("https://waimaieapp.meituan.com/bizdata/single/flow/origin", params=params, headers=headers)
    data = res.json()
    positions = jsonpath.jsonpath(data,"$..data.positions.*.position") 
    items = jsonpath.jsonpath(data,"$..data.positions.*.data")
    if items == False:
        return False
    result = {}
    for position, item in zip(positions, items):
        if item!= None:
            result[position] = {"exposeCnt":item["exposeCnt"], "visitCnt":item["visitCnt"]}

    params["tabType"] = 2 # 推广流量
    res = requests.get("https://waimaieapp.meituan.com/bizdata/single/flow/origin", params=params, headers=headers)
    data = res.json()    
    expCnts_ = jsonpath.jsonpath(data,"$..data.positions.*.data.exposeCnt")    
    visitCnts_ = jsonpath.jsonpath(data,"$..data.positions.*.data.visitCnt")
    if expCnts_ != False:
        result["推广"] = {"exposeCnt":sum(expCnts_), "visitCnt":sum(visitCnts_)}
    else:        
        result["推广"] = {"exposeCnt":0, "visitCnt":0}

    pos_list = ["商家列表", "其它", "搜索", "首页展位","推广"] # 数据汇总
    data_dic[shop_dic[shopId]] = [data_date]
    exp_list = [result[pos]["exposeCnt"] for pos in pos_list]
    data_dic[shop_dic[shopId]].append(sum(exp_list))
    data_dic[shop_dic[shopId]].extend(exp_list)      
    vst_list = [result[pos]["visitCnt"] for pos in pos_list]
    data_dic[shop_dic[shopId]].append(sum(vst_list))
    data_dic[shop_dic[shopId]].extend(vst_list)


data_dic, shop_dic ={}, {}

def main(token_dic):   

    global shop_dic
    shop_dic = token_dic["mt_shop_dic"]
    acctId = token_dic["acctId"]
    cookie = token_dic["cookie"]
    token = token_dic["token"]
    for shop in shop_dic:
        if fetch(shop, token, acctId, cookie) == False:
            return False
        
    path = "流量表.xlsx"
    if os.path.exists(path) == False:
        wb = openpyxl.load_workbook(r"..\templete\流量表模板.xlsx")
        ws = wb["美团"]
        for shopName in shop_dic.values():
            _ = wb.copy_worksheet(ws)
            _.title = shopName
        wb.remove(ws)
        wb.save(path)
        
    wb = openpyxl.load_workbook(path)            
    up_font = openpyxl.styles.Font(color="FF0000")   # 格式美化
    down_font = openpyxl.styles.Font(color="00FFFF")
    shops = shop_dic.values()
    for shop in shops:
        wb[shop].append(data_dic[shop]) # 写入数据
        max_row = wb[shop].max_row
        while wb[shop].cell(max_row,1).value is None:
            max_row -= 1
        for col in range(1,15):
            wb[shop].cell(max_row,col).alignment = openpyxl.styles.Alignment("center")
        if max_row >= 9 :
            for col in range(2,14):
                if wb[shop].cell(max_row,col).value >= wb[shop].cell(max_row-7,col).value * 1.05 :
                    wb[shop].cell(max_row,col).font = up_font
                    continue
                if wb[shop].cell(max_row,col).value <= wb[shop].cell(max_row-7,col).value * 0.95 :
                    wb[shop].cell(max_row,col).font = down_font                                          
    wb.save(path)
